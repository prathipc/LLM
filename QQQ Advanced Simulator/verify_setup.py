#!/usr/bin/env python3
"""
Verification script to test all components before running the full simulation
"""

import sys
import os

print("=" * 70)
print("QQQ ADVANCED SIMULATOR - SETUP VERIFICATION")
print("=" * 70)

# Change to script directory
os.chdir(os.path.dirname(os.path.abspath(__file__)))

# Test 1: Python version
print("\n1. Checking Python version...")
if sys.version_info < (3, 6):
    print("   ❌ ERROR: Python 3.6+ required")
    sys.exit(1)
print(f"   ✓ Python {sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}")

# Test 2: Required packages
print("\n2. Checking required packages...")
required = ['pandas', 'numpy', 'matplotlib', 'openpyxl']
missing = []
for pkg in required:
    try:
        __import__(pkg)
        print(f"   ✓ {pkg}")
    except ImportError:
        print(f"   ❌ {pkg} - NOT INSTALLED")
        missing.append(pkg)

if missing:
    print(f"\n   Install missing packages with:")
    print(f"   pip3 install {' '.join(missing)}")
    sys.exit(1)

# Test 3: Input file
print("\n3. Checking input file...")
input_file = "data/Input Files/QQQ.xlsx"
if not os.path.exists(input_file):
    print(f"   ❌ ERROR: {input_file} not found")
    sys.exit(1)
print(f"   ✓ {input_file} exists")

# Test 4: Load and verify data
print("\n4. Loading and verifying data...")
from openpyxl import load_workbook
import pandas as pd

try:
    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb['QQQ']
    headers = [cell.value for cell in ws[1]]
    
    required_cols = ['Date', 'QQQ Close Price', 'TQQQ Price', 'SQQQ Price']
    for col in required_cols:
        if col in headers:
            print(f"   ✓ Column '{col}' found")
        else:
            print(f"   ❌ Column '{col}' missing")
            sys.exit(1)
    
    # Count rows
    row_count = sum(1 for _ in ws.iter_rows(min_row=2))
    print(f"   ✓ Found {row_count} data rows")
    
    wb.close()
    
except Exception as e:
    print(f"   ❌ ERROR: {e}")
    sys.exit(1)

# Test 5: Import modules
print("\n5. Testing module imports...")
sys.path.insert(0, 'src')
try:
    from config import *
    print("   ✓ config.py")
    
    from indicators import calculate_all_indicators
    print("   ✓ indicators.py")
    
    from trade_engine import TradeEngine
    print("   ✓ trade_engine.py")
    
    from reporting import generate_all_reports
    print("   ✓ reporting.py")
    
except Exception as e:
    print(f"   ❌ ERROR: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Test 6: Output directory
print("\n6. Checking output directory...")
output_dir = "data/Output Files"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
    print(f"   ✓ Created {output_dir}")
else:
    print(f"   ✓ {output_dir} exists")

# Test 7: Quick simulation test
print("\n7. Running quick simulation test (100 days)...")
try:
    import time
    start = time.time()
    
    # Load data
    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb['QQQ']
    headers = [cell.value for cell in ws[1]]
    data = [row for row in ws.iter_rows(min_row=2, max_row=102, values_only=True)]
    df = pd.DataFrame(data, columns=headers)
    wb.close()
    
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.sort_values('Date').reset_index(drop=True)
    
    # Calculate indicators
    df = calculate_all_indicators(df)
    df = df.dropna()
    
    # Run mini simulation
    engine = TradeEngine(initial_capital=10000)
    for idx, row in df.iterrows():
        engine.process_day(row['Date'], row)
    
    elapsed = time.time() - start
    final_value = engine.get_portfolio_value(
        df.iloc[-1]['TQQQ Price'], 
        df.iloc[-1]['SQQQ Price']
    )
    
    print(f"   ✓ Test completed in {elapsed:.2f} seconds")
    print(f"   ✓ Portfolio value: ${final_value:.2f}")
    
except Exception as e:
    print(f"   ❌ ERROR: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# All tests passed!
print("\n" + "=" * 70)
print("✓ ALL TESTS PASSED!")
print("=" * 70)
print("\nYou can now run the full simulation:")
print("  python3 src/main_fixed.py")
print("=" * 70)

