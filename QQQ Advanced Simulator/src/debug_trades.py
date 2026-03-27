#!/usr/bin/env python3
"""
Debug script to understand why no trades were executed
"""

import os
import sys
import pandas as pd
from openpyxl import load_workbook

os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, '.')

from config import *
from indicators import calculate_all_indicators

# Load data
print("Loading data...")
wb = load_workbook('../data/Input Files/QQQ.xlsx', read_only=True, data_only=True)
ws = wb['QQQ']
headers = [cell.value for cell in ws[1]]
data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
df = pd.DataFrame(data, columns=headers)
wb.close()

df['Date'] = pd.to_datetime(df['Date'])
df = df.sort_values('Date').reset_index(drop=True)
df = df[df['Date'] >= '2021-01-01'].reset_index(drop=True)

# Calculate indicators
print("Calculating indicators...")
df = calculate_all_indicators(df)
df = df.dropna()

print(f"\nAnalyzing {len(df)} trading days from {df['Date'].min()} to {df['Date'].max()}")
print("=" * 80)

# Check key conditions for 100% TQQQ entry (Rule 3)
print("\n100% TQQQ Entry Conditions (Rule 3):")
print("-" * 80)

rule3_price_above_50 = df[COL_CLOSE] > df['SMA_50'] * (1 + BUFFER_ENTRY_50SMA)
rule3_price_above_200 = df[COL_CLOSE] > df['SMA_200'] * (1 + BUFFER_ENTRY_200SMA)
rule3_sma_golden = df['SMA_50'] > df['SMA_200']
rule3_momentum = df['Combined_Momentum'] > MOMENTUM_VERY_STRONG_POSITIVE
rule3_slope_50 = df['Slope_50'] > SLOPE_BULL_THRESHOLD
rule3_slope_200 = df['Slope_200'] > 0
rule3_volume = df['Volume_Ratio'] >= VOLUME_RATIO_HIGH

print(f"1. Price > 50-SMA * 1.015:        {rule3_price_above_50.sum():4d} days ({rule3_price_above_50.mean()*100:.1f}%)")
print(f"2. Price > 200-SMA * 1.02:        {rule3_price_above_200.sum():4d} days ({rule3_price_above_200.mean()*100:.1f}%)")
print(f"3. SMA50 > SMA200 (Golden Cross): {rule3_sma_golden.sum():4d} days ({rule3_sma_golden.mean()*100:.1f}%)")
print(f"4. Momentum > {MOMENTUM_VERY_STRONG_POSITIVE}:              {rule3_momentum.sum():4d} days ({rule3_momentum.mean()*100:.1f}%)")
print(f"5. Slope_50 > {SLOPE_BULL_THRESHOLD}:                {rule3_slope_50.sum():4d} days ({rule3_slope_50.mean()*100:.1f}%)")
print(f"6. Slope_200 > 0:                 {rule3_slope_200.sum():4d} days ({rule3_slope_200.mean()*100:.1f}%)")
print(f"7. Volume Ratio >= {VOLUME_RATIO_HIGH}:            {rule3_volume.sum():4d} days ({rule3_volume.mean()*100:.1f}%)")

rule3_all = (rule3_price_above_50 & rule3_price_above_200 & rule3_sma_golden & 
             rule3_momentum & rule3_slope_50 & rule3_slope_200 & rule3_volume)
print(f"\nALL CONDITIONS MET:               {rule3_all.sum():4d} days ({rule3_all.mean()*100:.1f}%)")

if rule3_all.sum() > 0:
    print("\nDates when ALL Rule 3 conditions were met:")
    print(df[rule3_all][['Date', COL_CLOSE, 'SMA_50', 'SMA_200', 'Combined_Momentum', 'Slope_50', 'Volume_Ratio']].head(10))

# Check Rule 4 (75% TQQQ)
print("\n\n75% TQQQ Entry Conditions (Rule 4):")
print("-" * 80)

rule4_price_above_50 = df[COL_CLOSE] > df['SMA_50'] * (1 + BUFFER_ENTRY_50SMA)
rule4_price_above_200 = df[COL_CLOSE] > df['SMA_200']
rule4_sma_golden = df['SMA_50'] > df['SMA_200']
rule4_momentum = (df['Combined_Momentum'] >= MOMENTUM_MODERATE_POSITIVE) & (df['Combined_Momentum'] <= MOMENTUM_VERY_STRONG_POSITIVE)
rule4_volume = df['Volume_Ratio'] >= VOLUME_RATIO_LOW

print(f"1. Price > 50-SMA * 1.015:         {rule4_price_above_50.sum():4d} days ({rule4_price_above_50.mean()*100:.1f}%)")
print(f"2. Price > 200-SMA:                {rule4_price_above_200.sum():4d} days ({rule4_price_above_200.mean()*100:.1f}%)")
print(f"3. SMA50 > SMA200:                 {rule4_sma_golden.sum():4d} days ({rule4_sma_golden.mean()*100:.1f}%)")
print(f"4. Momentum {MOMENTUM_MODERATE_POSITIVE} to {MOMENTUM_VERY_STRONG_POSITIVE}:         {rule4_momentum.sum():4d} days ({rule4_momentum.mean()*100:.1f}%)")
print(f"5. Volume Ratio >= {VOLUME_RATIO_LOW}:             {rule4_volume.sum():4d} days ({rule4_volume.mean()*100:.1f}%)")

rule4_all = rule4_price_above_50 & rule4_price_above_200 & rule4_sma_golden & rule4_momentum & rule4_volume
print(f"\nALL CONDITIONS MET:                {rule4_all.sum():4d} days ({rule4_all.mean()*100:.1f}%)")

if rule4_all.sum() > 0:
    print("\nDates when ALL Rule 4 conditions were met:")
    print(df[rule4_all][['Date', COL_CLOSE, 'SMA_50', 'SMA_200', 'Combined_Momentum', 'Volume_Ratio']].head(10))

# Check Rule 5 (50% TQQQ)
print("\n\n50% TQQQ Entry Conditions (Rule 5):")
print("-" * 80)

rule5_price_above_200 = df[COL_CLOSE] > df['SMA_200']
rule5_price_near_50 = df[COL_CLOSE] > df['SMA_50'] * 0.98
rule5_momentum = (df['Combined_Momentum'] >= MOMENTUM_NEUTRAL_LOW) & (df['Combined_Momentum'] <= MOMENTUM_MODERATE_POSITIVE)

print(f"1. Price > 200-SMA:                {rule5_price_above_200.sum():4d} days ({rule5_price_above_200.mean()*100:.1f}%)")
print(f"2. Price > 50-SMA * 0.98:          {rule5_price_near_50.sum():4d} days ({rule5_price_near_50.mean()*100:.1f}%)")
print(f"3. Momentum {MOMENTUM_NEUTRAL_LOW} to {MOMENTUM_MODERATE_POSITIVE}:    {rule5_momentum.sum():4d} days ({rule5_momentum.mean()*100:.1f}%)")

rule5_all = rule5_price_above_200 & rule5_price_near_50 & rule5_momentum
print(f"\nALL CONDITIONS MET:                {rule5_all.sum():4d} days ({rule5_all.mean()*100:.1f}%)")

if rule5_all.sum() > 0:
    print("\nDates when ALL Rule 5 conditions were met:")
    print(df[rule5_all][['Date', COL_CLOSE, 'SMA_50', 'SMA_200', 'Combined_Momentum']].head(10))

# Summary statistics
print("\n\n" + "=" * 80)
print("SUMMARY STATISTICS")
print("=" * 80)
print(f"Combined Momentum - Min: {df['Combined_Momentum'].min():.2f}, Max: {df['Combined_Momentum'].max():.2f}, Mean: {df['Combined_Momentum'].mean():.2f}")
print(f"Slope 50          - Min: {df['Slope_50'].min():.2f}, Max: {df['Slope_50'].max():.2f}, Mean: {df['Slope_50'].mean():.2f}")
print(f"Slope 200         - Min: {df['Slope_200'].min():.2f}, Max: {df['Slope_200'].max():.2f}, Mean: {df['Slope_200'].mean():.2f}")
print(f"Volume Ratio      - Min: {df['Volume_Ratio'].min():.2f}, Max: {df['Volume_Ratio'].max():.2f}, Mean: {df['Volume_Ratio'].mean():.2f}")
print(f"ATR %             - Min: {df['ATR_PCT'].min():.2f}, Max: {df['ATR_PCT'].max():.2f}, Mean: {df['ATR_PCT'].mean():.2f}")

print(f"\nGolden Cross (SMA50 > SMA200): {(df['SMA_50'] > df['SMA_200']).sum()} days ({(df['SMA_50'] > df['SMA_200']).mean()*100:.1f}%)")
print(f"Death Cross (SMA50 < SMA200):  {(df['SMA_50'] < df['SMA_200']).sum()} days ({(df['SMA_50'] < df['SMA_200']).mean()*100:.1f}%)")


