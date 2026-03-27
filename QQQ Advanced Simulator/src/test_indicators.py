"""
Test indicator calculations to find bottleneck
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import time

# Import config
from config import *

print("=" * 70)
print("TESTING INDICATOR CALCULATIONS")
print("=" * 70)

# Load data
print("\n1. Loading data...")
start = time.time()
wb = load_workbook("data/Input Files/QQQ.xlsx", read_only=True, data_only=True)
ws = wb['QQQ']
headers = [cell.value for cell in ws[1]]
data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
df = pd.DataFrame(data, columns=headers)
wb.close()
df['Date'] = pd.to_datetime(df['Date'])
df = df.sort_values('Date').reset_index(drop=True)
df = df[df['Date'] >= '2021-01-01'].reset_index(drop=True)
print(f"   Loaded {len(df)} rows in {time.time()-start:.2f}s")

# Test each indicator function
print("\n2. Testing SMA 50...")
start = time.time()
df['SMA_50'] = df['QQQ Close Price'].rolling(window=50, min_periods=50).mean()
print(f"   Done in {time.time()-start:.2f}s")

print("\n3. Testing SMA 200...")
start = time.time()
df['SMA_200'] = df['QQQ Close Price'].rolling(window=200, min_periods=200).mean()
print(f"   Done in {time.time()-start:.2f}s")

print("\n4. Testing SMA 50 slope...")
start = time.time()
from indicators import calculate_sma_slope
df['Slope_50'] = calculate_sma_slope(df['SMA_50'])
print(f"   Done in {time.time()-start:.2f}s")

print("\n5. Testing SMA 200 slope...")
start = time.time()
df['Slope_200'] = calculate_sma_slope(df['SMA_200'])
print(f"   Done in {time.time()-start:.2f}s")

print("\n6. Testing ATR...")
start = time.time()
from indicators import calculate_atr, calculate_atr_percentage
df['ATR'] = calculate_atr(df['QQQ High Price'], df['QQQ Low Price'], df['QQQ Close Price'])
df['ATR_PCT'] = calculate_atr_percentage(df['ATR'], df['QQQ Close Price'])
print(f"   Done in {time.time()-start:.2f}s")

print("\n7. Testing Volume Ratio...")
start = time.time()
from indicators import calculate_volume_ratio
df['Volume_Ratio'] = calculate_volume_ratio(df['Volume'])
print(f"   Done in {time.time()-start:.2f}s")

print("\n8. Testing Combined Momentum...")
start = time.time()
from indicators import calculate_combined_momentum
df['Combined_Momentum'] = calculate_combined_momentum(
    df['Slope_50'], df['Slope_200'], df['QQQ Close Price'], df['SMA_50'], df['SMA_200']
)
print(f"   Done in {time.time()-start:.2f}s")

print("\n9. Testing Market Regime...")
start = time.time()
from indicators import detect_market_regime
df['Market_Regime'] = detect_market_regime(
    df['QQQ Close Price'], df['SMA_50'], df['SMA_200'],
    df['Slope_50'], df['Slope_200'], df['Volume_Ratio'], df['ATR_PCT']
)
print(f"   Done in {time.time()-start:.2f}s")

print("\n10. Testing calculate_all_indicators function...")
start = time.time()
from indicators import calculate_all_indicators

# Reload fresh data
wb = load_workbook("data/Input Files/QQQ.xlsx", read_only=True, data_only=True)
ws = wb['QQQ']
headers = [cell.value for cell in ws[1]]
data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
df2 = pd.DataFrame(data, columns=headers)
wb.close()
df2['Date'] = pd.to_datetime(df2['Date'])
df2 = df2.sort_values('Date').reset_index(drop=True)
df2 = df2[df2['Date'] >= '2021-01-01'].reset_index(drop=True)

print("   Calling calculate_all_indicators...")
df2 = calculate_all_indicators(df2)
print(f"   Done in {time.time()-start:.2f}s")

print("\n" + "=" * 70)
print("ALL TESTS PASSED!")
print("=" * 70)

