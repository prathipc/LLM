"""
Minimal debug version to identify bottlenecks
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
import time

def time_section(name):
    """Decorator to time function execution"""
    def decorator(func):
        def wrapper(*args, **kwargs):
            print(f"\n>>> Starting: {name}")
            start = time.time()
            result = func(*args, **kwargs)
            elapsed = time.time() - start
            print(f"<<< Completed: {name} in {elapsed:.2f} seconds")
            return result
        return wrapper
    return decorator


@time_section("Load Excel Data")
def load_data():
    """Load data from Excel"""
    file_path = "data/Input Files/QQQ.xlsx"
    
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['QQQ']
    
    headers = [cell.value for cell in ws[1]]
    print(f"  Headers: {headers}")
    
    data = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        data.append(row)
        if i % 100 == 0:
            print(f"  Reading row {i}...", end='\r')
    
    print(f"  Total rows read: {len(data)}")
    
    df = pd.DataFrame(data, columns=headers)
    wb.close()
    
    return df


@time_section("Process Data")
def process_data(df):
    """Basic data processing"""
    
    # Convert date
    print("  Converting dates...")
    df['Date'] = pd.to_datetime(df['Date'])
    
    # Sort
    print("  Sorting...")
    df = df.sort_values('Date').reset_index(drop=True)
    
    # Filter 2021+
    print("  Filtering for 2021+...")
    df = df[df['Date'] >= '2021-01-01'].reset_index(drop=True)
    print(f"  Rows after filter: {len(df)}")
    
    return df


@time_section("Calculate SMA 50")
def calc_sma_50(df):
    """Calculate 50-day SMA"""
    df['SMA_50'] = df['QQQ Close Price'].rolling(window=50, min_periods=50).mean()
    return df


@time_section("Calculate SMA 200")
def calc_sma_200(df):
    """Calculate 200-day SMA"""
    df['SMA_200'] = df['QQQ Close Price'].rolling(window=200, min_periods=200).mean()
    return df


@time_section("Calculate Slopes")
def calc_slopes(df):
    """Calculate SMA slopes"""
    lookback = 5
    
    # 50 SMA slope
    sma_n_ago = df['SMA_50'].shift(lookback)
    mask = sma_n_ago > 0
    slope = pd.Series(0.0, index=df.index)
    slope[mask] = ((df['SMA_50'][mask] - sma_n_ago[mask]) / sma_n_ago[mask]) / lookback * 100 * 5
    df['Slope_50'] = slope
    
    # 200 SMA slope
    sma_n_ago = df['SMA_200'].shift(lookback)
    mask = sma_n_ago > 0
    slope = pd.Series(0.0, index=df.index)
    slope[mask] = ((df['SMA_200'][mask] - sma_n_ago[mask]) / sma_n_ago[mask]) / lookback * 100 * 5
    df['Slope_200'] = slope
    
    return df


@time_section("Simple Trading Simulation")
def simple_simulation(df):
    """Very simple trading simulation"""
    
    # Remove warm-up period
    df = df.dropna()
    print(f"  Rows after dropna: {len(df)}")
    
    cash = 10000
    position = 'CASH'
    portfolio_values = []
    
    for i, row in df.iterrows():
        if i % 50 == 0:
            print(f"  Processing day {i}/{len(df)}...", end='\r')
        
        # Very simple logic: buy TQQQ if price > SMA_50
        price_qqq = row['QQQ Close Price']
        sma_50 = row['SMA_50']
        
        if position == 'CASH' and price_qqq > sma_50:
            # Buy TQQQ
            shares = cash / row['TQQQ Price']
            position = 'TQQQ'
            position_value = shares * row['TQQQ Price']
            portfolio_values.append(position_value)
        elif position == 'TQQQ' and price_qqq < sma_50:
            # Sell TQQQ
            shares = cash / row['TQQQ Price']
            cash = shares * row['TQQQ Price']
            position = 'CASH'
            portfolio_values.append(cash)
        elif position == 'TQQQ':
            shares = cash / row['TQQQ Price']
            position_value = shares * row['TQQQ Price']
            portfolio_values.append(position_value)
        else:
            portfolio_values.append(cash)
    
    print(f"\n  Final portfolio value: ${portfolio_values[-1]:.2f}")
    return portfolio_values


def main():
    print("=" * 70)
    print("DEBUG RUN - MINIMAL VERSION")
    print("=" * 70)
    
    overall_start = time.time()
    
    # Load
    df = load_data()
    
    # Process
    df = process_data(df)
    
    # Calculate indicators
    df = calc_sma_50(df)
    df = calc_sma_200(df)
    df = calc_slopes(df)
    
    # Simulate
    values = simple_simulation(df)
    
    overall_elapsed = time.time() - overall_start
    
    print("\n" + "=" * 70)
    print(f"TOTAL TIME: {overall_elapsed:.2f} seconds")
    print("=" * 70)


if __name__ == "__main__":
    main()

