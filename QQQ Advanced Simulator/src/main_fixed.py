"""
Main Script for Enhanced Leveraged ETF Strategy v2.0
Starting from Jan 2021 with performance optimizations
"""

print("="*70)
print("SCRIPT STARTING...")
print("="*70)
import sys
sys.stdout.flush()

import pandas as pd
import numpy as np
import os
import sys
from datetime import datetime
from openpyxl import load_workbook

# Import modules
from config import *
from indicators import calculate_all_indicators
from trade_engine import TradeEngine
from reporting import generate_all_reports


def load_data_optimized(file_path):
    """
    Load price data from Excel file efficiently
    Uses openpyxl directly for better performance
    """
    print("=" * 70)
    print("LOADING DATA (FROM JAN 2021)")
    print("=" * 70)
    
    try:
        # Adjust path if running from src directory
        import os
        if not os.path.exists(file_path):
            file_path = file_path.replace('../', '')
        
        print(f"Reading Excel file: {file_path}")
        
        # Load workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb['QQQ']
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"  Headers: {headers}")
        
        # Read all data
        print("  Reading all rows...")
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        print(f"  Loaded {len(df)} total rows")
        
        # Close workbook
        wb.close()
        
        # Convert Date column to datetime
        df[COL_DATE] = pd.to_datetime(df[COL_DATE])
        
        # Sort by date ascending
        df = df.sort_values(COL_DATE).reset_index(drop=True)
        print(f"  Date range: {df[COL_DATE].min()} to {df[COL_DATE].max()}")
        
        # Filter for 2021 onwards
        start_date = pd.Timestamp('2021-01-01')
        df = df[df[COL_DATE] >= start_date].reset_index(drop=True)
        print(f"  After filtering (2021+): {len(df)} rows")
        print(f"  Filtered date range: {df[COL_DATE].min()} to {df[COL_DATE].max()}")
        
        # Verify required columns
        required_cols = [COL_DATE, COL_CLOSE, COL_OPEN, COL_HIGH, COL_LOW, 
                        COL_VOLUME, COL_TQQQ_CLOSE, COL_SQQQ_CLOSE]
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            print(f"  ERROR: Missing columns: {missing_cols}")
            print(f"  Available columns: {df.columns.tolist()}")
            sys.exit(1)
        
        # Check for nulls
        null_counts = df[required_cols].isnull().sum()
        if null_counts.any():
            print("\n  Warning: Null values found:")
            print(null_counts[null_counts > 0])
        
        # Forward fill any missing leveraged ETF prices
        df[COL_TQQQ_CLOSE] = df[COL_TQQQ_CLOSE].ffill()
        df[COL_SQQQ_CLOSE] = df[COL_SQQQ_CLOSE].ffill()
        
        # Drop any remaining rows with nulls in critical columns
        critical_cols = [COL_DATE, COL_CLOSE, COL_TQQQ_CLOSE, COL_SQQQ_CLOSE]
        df = df.dropna(subset=critical_cols)
        print(f"  After removing nulls: {len(df)} rows")
        
        print("\nData sample:")
        print(df.head(3))
        print("...")
        print(df.tail(3))
        
        return df
        
    except Exception as e:
        print(f"\nERROR loading data: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def main():
    """Main execution function"""
    print("\n" + "=" * 70)
    print("ENHANCED LEVERAGED ETF STRATEGY v2.0")
    print("=" * 70)
    print(f"Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    # Ensure output directory exists
    output_dir_abs = os.path.join(os.path.dirname(__file__), '..', OUTPUT_DIR)
    output_dir_abs = os.path.abspath(output_dir_abs)
    os.makedirs(output_dir_abs, exist_ok=True)
    print(f"Output directory: {output_dir_abs}\n")
    
    # Load data
    data = load_data_optimized(INPUT_FILE_PATH)
    
    if len(data) < 250:
        print(f"\nERROR: Insufficient data ({len(data)} rows). Need at least 250 days.")
        return
    
    # Calculate indicators
    print("\n" + "=" * 70)
    print("CALCULATING INDICATORS")
    print("=" * 70)
    
    start_time = datetime.now()
    data = calculate_all_indicators(data)
    elapsed = (datetime.now() - start_time).total_seconds()
    print(f"  Indicators calculated in {elapsed:.2f} seconds")
    
    # Drop rows with NaN (from indicator calculation warm-up)
    initial_rows = len(data)
    data = data.dropna()
    print(f"  Dropped {initial_rows - len(data)} warm-up rows, {len(data)} rows remaining")
    
    # Run simulation
    print("\n" + "=" * 70)
    print("RUNNING TRADING SIMULATION")
    print("=" * 70)
    print(f"  Initial Capital: ${INITIAL_CAPITAL:,.2f}")
    print(f"  Trading Days: {len(data)}")
    print()
    
    engine = TradeEngine(initial_capital=INITIAL_CAPITAL)
    
    start_time = datetime.now()
    total_days = len(data)
    
    # Process each day with progress indicator
    for idx, row in data.iterrows():
        # Progress indicator every 50 days
        if idx % 50 == 0:
            progress = (idx / total_days) * 100
            elapsed = (datetime.now() - start_time).total_seconds()
            print(f"  Progress: {progress:.1f}% ({idx}/{total_days} days, {elapsed:.1f}s elapsed)")
        
        engine.process_day(row[COL_DATE], row)
    
    elapsed = (datetime.now() - start_time).total_seconds()
    print(f"\n  Simulation completed in {elapsed:.2f} seconds")
    last_row = data.iloc[-1]
    print(f"  Final Portfolio Value: ${engine.get_portfolio_value(last_row[COL_TQQQ_CLOSE], last_row[COL_SQQQ_CLOSE]):.2f}")
    
    # Get results
    ledger_df = pd.DataFrame(engine.daily_ledger)
    trades_df = pd.DataFrame(engine.trades)
    
    print(f"\n  Total Trades: {len(trades_df)}")
    engine.print_signal_diagnostics()
    
    # Generate reports
    print("\n" + "=" * 70)
    print("GENERATING REPORTS")
    print("=" * 70)
    
    start_time = datetime.now()
    output_dir_abs = os.path.join(os.path.dirname(__file__), '..', OUTPUT_DIR)
    output_dir_abs = os.path.abspath(output_dir_abs)
    generate_all_reports(ledger_df, trades_df, INITIAL_CAPITAL, output_dir_abs)
    elapsed = (datetime.now() - start_time).total_seconds()
    print(f"  Reports generated in {elapsed:.2f} seconds")
    
    print("\n" + "=" * 70)
    print("SIMULATION COMPLETE!")
    print("=" * 70)
    print(f"End Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"\nAll outputs saved to: {output_dir_abs}")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()

